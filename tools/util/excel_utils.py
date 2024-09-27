from openpyxl import load_workbook
from util import file_utils
from collections import Counter


def rename_by_data(file_path):
    """
    根据Excel文件中的数据重新命名文件。

    加载Excel文件并根据其内部数据生成新的文件名。

    参数:
    file_path (str): 原始Excel文件的路径。

    无返回值。
    """
    wb = load_workbook(filename=file_utils.xls_to_xlsx(file_path))
    ws = wb.active

    # 初始化新的文件名为空字符串
    new_filename = ""
    # 遍历工作表中的所有列
    for col in ws.iter_cols(values_only=True):
        # 统计当前列各个值的出现次数
        col_values_counter = Counter(col)
        # 如果列中不同值的数量为2时（只有列名和一个值）
        if len(col_values_counter) == 2:
            # 第二个值（即实际数据）不能为None或者空值
            if col[1] != None and col[1] != "":
                key = str(col[0])
                val = str(col[1])
                new_filename += f"{key}-{val}_"

    # 获取文件的创建时间，并格式化
    formatted_time = file_utils.get_file_created_time(file_path)
    # 最终文件名由新文件名基础部分和创建时间组成，加上文件扩展名
    final_name = new_filename + formatted_time + ".xlsx"

    # 重命名文件
    file_utils.rename_file(file_path, final_name)